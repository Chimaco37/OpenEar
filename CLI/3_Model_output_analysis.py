def polygon_to_bbox(polygon):
    x_coords, y_coords = zip(*polygon)
    min_x, max_x = min(x_coords), max(x_coords)
    min_y, max_y = min(y_coords), max(y_coords)
    w = round(max_x - min_x, 7)
    h = round(max_y - min_y, 7)

    polygon = np.array(polygon, dtype=np.float32)
    area = Polygon(polygon).area 

    return min_x, min_y, max_x, max_y, w, h, area

def pre_process_projection_label_file(projection_label_file_path):
    with open(projection_label_file_path, 'r') as infile:
        lines = infile.read().splitlines()

    bounding_boxes_data = []
    i = 0
    for line in lines:
        try:
            data = line.split()
            category = data[0]
            polygon = []
            for i in range(1, len(data), 2):
                x = float(data[i])
                y = float(data[i + 1])
                polygon.append((x, y))
            bbox = tuple(polygon_to_bbox(polygon))
            i = i + 1

            bounding_boxes_data.append(bbox)
        except:
            i = i + 1
            continue

    return bounding_boxes_data

def pre_process_ear_label_file(ear_file_path):
    polygon = []
    with open(ear_file_path, 'r') as file:
        data = file.readline().split()
        category = data[0]
        for i in range(1, len(data), 2):
            x = float(data[i]) * image_width
            y = float(data[i + 1]) * image_height
            polygon.append((x, y))

    polygon = np.array(polygon, dtype=np.float32)
    return polygon

#################################################### kernel traits #####################################################
def calculate_kernel_number(bboxes_data, start_index, end_index):
    kernel_number = 0.0

    for bbox in bboxes_data:
        min_x = bbox[0]
        max_x = bbox[2]
        w = bbox[4] # width
        
        if w <= 0:
            continue
            
        overlap_start = max(min_x, start_index)
        overlap_end = min(max_x, end_index)

        overlap_len = max(0.0, overlap_end - overlap_start)

        ratio = overlap_len / w
        kernel_number += ratio

    return int(kernel_number)

# def calculate_kernel_number(bboxes_data, start_index, end_index):
#     bboxes_data = [bbox for bbox in bboxes_data if
#                    start_index < (bbox[0] + bbox[2]) / 2 < end_index]

#     kernel_number = len(bboxes_data)

#     return kernel_number

def calculate_kernel_row_number(bboxes_data, start_index, end_index):
    if start_index >= end_index:
        start_index, end_index = 0.5, 0.5

    bboxes_data = filter_regular_kernels(bboxes_data, start_index, end_index)

    if not bboxes_data:
        return 'NA'

    ymin_values = [bbox[1] for bbox in bboxes_data]
    ymax_values = [bbox[3] for bbox in bboxes_data]

    LOW = min(ymin_values)
    HIGH = max(ymax_values)

    num_intervals = 200
    step_size = (HIGH - LOW) / num_intervals

    row_counts = []
    threshold = LOW

    for i in range(num_intervals):
        threshold = LOW + i * step_size
        row_bbox_count = sum(1 for bbox in bboxes_data if bbox[1] < threshold < bbox[3])

        raw_row_count = (row_bbox_count + 1) // 2 * 2  
        row_counts.append(raw_row_count)

    kernel_row_number = round(max(set(row_counts), key=row_counts.count),0)

    return kernel_row_number


def calculate_kernel_temp_value(bboxes_data, start_index, end_index):

    bboxes_data = filter_regular_kernels(bboxes_data, start_index, end_index)

    if not bboxes_data:
        return ''

    ymin_values = [bbox[1] for bbox in bboxes_data]
    ymax_values = [bbox[3] for bbox in bboxes_data]

    LOW = min(ymin_values)
    HIGH = max(ymax_values)

    kernels_area_height = HIGH - LOW
    inside_kernel_number = len(bboxes_data)

    kernel_temp_value = kernels_area_height / inside_kernel_number

    return kernel_temp_value

def calculate_thousand_kernel_area_temp_value(bboxes_data, start_index, end_index):

    filtered_bboxes_data = filter_regular_kernels(bboxes_data, start_index, end_index)

    if not bboxes_data:
        return ''

    area_sum = 0

    for bbox in filtered_bboxes_data:
        area_sum += bbox[6]

    filtered_kernel_number = len(bboxes_data)

    thousand_kernel_area_temp_value = area_sum / filtered_kernel_number * 1000

    return thousand_kernel_area_temp_value

def filter_regular_kernels(bboxes_data, start_index, end_index):
    bboxes_data = [
        bbox for bbox in bboxes_data
        if start_index < (bbox[0] + bbox[2]) / 2 < end_index
    ]

    bboxes_data.sort(key=lambda x: x[3])

    v_start_index = int(len(bboxes_data) / 16)
    v_end_index = int(len(bboxes_data) * 15 / 16)

    filtered_bboxes = bboxes_data[v_start_index:v_end_index]

    return filtered_bboxes

#################################################### Ear traits ########################################################
def calculate_ear_length_diameter(polygon):
    (cx, cy), (l, w), theta = cv2.minAreaRect(polygon)

    polygon = Polygon(polygon) 

    cutting_line = create_cutting_line(cx, cy, theta, l, w)

    intersection_line = cutting_line.intersection(polygon)

    ear_length = max(l,w) * scale_ratio
    ear_diameter = intersection_line.length * scale_ratio

    return ear_length, ear_diameter

def calculate_ear_volume(polygon):
    distances, interval_h = find_intersection_distances(polygon)
    h = interval_h * scale_ratio
    ear_volume = 0
    for distance in distances:
        radius = distance / 2 * scale_ratio
        ear_volume += math.pi * radius ** 2 * h

    return ear_volume


def create_cutting_line(cx, cy, theta, l, w):

    if l >= w:
        rad = np.deg2rad(-theta)
        dx = np.sin(rad) * MAX_LEN  
        dy = np.cos(rad) * MAX_LEN
    else:
        rad = np.deg2rad(theta)  
        dx = np.cos(rad) * MAX_LEN
        dy = np.sin(rad) * MAX_LEN

    line = LineString([(cx - dx, cy - dy), (cx + dx, cy + dy)])
    return line


def find_intersection_distances(polygon):

    num_lines = 20
    intersection_distances = []
    _, min_y, _, max_y, _, h, _ = polygon_to_bbox(polygon)
    interval_h = h / num_lines
    polygon = Polygon(polygon)  

    for i in range(0, num_lines):
        y = i * interval_h + min_y
        line = LineString([(0, y), (image_width, y)])

        intersection_line = line.intersection(polygon)

        if intersection_line.is_empty:
            continue

        if intersection_line.geom_type == 'LineString':
            intersection_distance = intersection_line.length
            intersection_distances.append(intersection_distance)
        elif intersection_line.geom_type == 'MultiLineString':
            for part in intersection_line.geoms:
                intersection_distance = part.length
                intersection_distances.append(intersection_distance)

    return intersection_distances, interval_h

def average(data_list):
    if len(data_list) == 0:
        return None
    return sum(data_list) / len(data_list)

def process_projection_file(file_path, start_index, end_index):
    bounding_boxes_data = pre_process_projection_label_file(file_path)

    kernel_number = calculate_kernel_number(bounding_boxes_data, start_index, end_index)

    kernel_row_number = calculate_kernel_row_number(bounding_boxes_data, start_index + 0.035, end_index - 0.035)

    try:
        kernel_number_per_row = kernel_number // kernel_row_number
    except:
        kernel_number_per_row = "NA"

    kernel_temp_value = calculate_kernel_temp_value(bounding_boxes_data, start_index, end_index)
    try:
        kernel_thickness = kernel_temp_value * kernel_row_number * scale_ratio * image_height
    except:
        kernel_thickness = 'NA'

    thousand_kernel_area_temp_value = calculate_thousand_kernel_area_temp_value(bounding_boxes_data, start_index, end_index)

    return kernel_number, kernel_row_number, kernel_number_per_row, kernel_thickness, thousand_kernel_area_temp_value


def process_ear_file(file_path):
    polygon = pre_process_ear_label_file(file_path)

    ear_length, ear_diameter = calculate_ear_length_diameter(polygon)

    ear_volume = calculate_ear_volume(polygon)

    return ear_length, ear_diameter, ear_volume

def preprocess_image(image_path, target_size):
    image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)  
    image = cv2.resize(image, target_size)  
    image = np.expand_dims(image, axis=-1) 
    image = np.expand_dims(image, axis=0)  
    return image

def kernel_row_number_classification(image_path, model):
    image = preprocess_image(image_path, (640, 640))
    prediction = np.argmax(model.predict(image)) 
    return prediction

def phenotype_classification(image_path, model):
    image = preprocess_image(image_path, (640, 640))
    prediction = np.argmax(model.predict(image)) 
    return prediction

def save_results_to_json(label, start_horizontal_index, end_horizontal_index, kernel_row_number, visualize_path):
    output_path = os.path.join(visualize_path, f"{label}.json")
    height_data = {
        "start_index": start_horizontal_index,
        "end_index": end_horizontal_index,
        "kernel_row": kernel_row_number, 
        "data_availability": "all"
    }

    try:
        with open(output_path, 'w') as json_file:
            json.dump(height_data, json_file, indent=4)
    except:
        return None

import os
import cv2
from openpyxl import Workbook, load_workbook
import glob
import numpy as np
import threading
import tkinter as tk
import json
from shapely.geometry import LineString, Polygon
import math
import argparse
import shutil
from tensorflow.keras.models import load_model

scale_ratio = 2 / 100
image_height=1440
image_width=1072
projection_image_width = 1920
KN_range = 0.91
# KRN_range = 0.84
MAX_LEN = math.hypot(image_width, image_height)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Analyze ear and projection model output data")

    parser.add_argument('-i', '--projection_image_folder', default='./images/projection/', type=str, required=False, help='Path to the projection image folder')
    parser.add_argument('-e', '--ear_label_folder', default='./output/ear/labels/', type=str, required=False, help='Path to the ear label folder')
    parser.add_argument('-p', '--projection_label_folder', default='./output/projection/labels/', type=str, required=False, help='Path to the projection label folder')
    parser.add_argument('-o', '--output_path', default='./output/', type=str, required=False, help='Output file path')
    parser.add_argument('-m', '--model_path', default='./models/', type=str, required=False, help='CNN model path')

    args = parser.parse_args()

    projection_label_files = sorted(glob.glob(os.path.join(args.projection_label_folder, '*.txt')))

    visualize_path = os.path.join(args.output_path, 'visualize')
    os.makedirs(visualize_path, exist_ok=True)

    output_file = os.path.join(args.output_path, "ear_phenotyping.xlsx")
    if os.path.exists(output_file):
        shutil.os.remove(output_file)

    wb = Workbook()
    ws = wb.active
    ws.append(
        ["Labels", "Ear_Length", "Ear_Diameter", "Ear_Volume", "Ear_Weight",
         "Kernel_Number", "Kernel_Row_Number", "Kernel_Number_per_Row", 
         "Kernel_Thickness", "Kernel_Width", "Thousand_Kernel_Weight"])

    start_index = (1 - KN_range) / 2
    end_index = (1 + KN_range) / 2

    for projection_label_file in projection_label_files:
        label = os.path.basename(projection_label_file).split(".")[0]  
        kernel_number, kernel_row_number, kernel_number_per_row, kernel_thickness, thousand_kernel_area_temp_value = process_projection_file(projection_label_file, start_index, end_index)

        ear_label_files = glob.glob(os.path.join(args.ear_label_folder, f'{label}_*.txt'))
        ear_lengths, ear_diameters, ear_volumes = [], [], []
        for ear_label_file in ear_label_files:
            try:
                ear_length, ear_diameter, ear_volume = process_ear_file(ear_label_file)
                ear_lengths.append(ear_length)
                ear_diameters.append(ear_diameter)
                ear_volumes.append(ear_volume)
            except Exception as e:
                print(f"Error on {label}")
                print(f"Reason: {e}")

        ear_length = round(np.median(np.array(ear_lengths)), 2)
        ear_diameter = round(np.median(np.array(ear_diameters)), 2)
        ear_volume = round(np.median(np.array(ear_volumes)), 2)

        try:
            x_scale_ratio = ear_diameter * math.pi / (end_index - start_index) * projection_image_width
            thousand_kernel_area = thousand_kernel_area_temp_value * x_scale_ratio * scale_ratio
            thousand_kernel_weight = round((thousand_kernel_area * 0.58 + 66.84), 2)
        except:
            thousand_kernel_weight = 'NA'

        try:
            ear_weight = round((ear_volume * 0.65 + 35.86), 2)
        except:
            ear_weight = "NA"

        try:
            kernel_width = round(ear_diameter * math.pi / kernel_row_number, 3)
        except:
            kernel_width = 'NA'

        try:
            kernel_thickness = round(kernel_thickness, 3)
        except:
            kernel_thickness = 'NA'
            
        save_results_to_json(label, start_index, end_index, kernel_row_number, visualize_path)

        ws.append([label, ear_length, ear_diameter, ear_volume, ear_weight, kernel_number, kernel_row_number,
                   kernel_number_per_row, kernel_thickness, kernel_width, thousand_kernel_weight])

    wb.save(output_file)

    try:
        model1 = load_model(os.path.join(args.model_path, "1_Developmental_Status_Assesment.h5"))
        model2 = load_model(os.path.join(args.model_path, "2_Kernel_Row_Visibility_Assesment.h5"))
    except Exception as e:
        print(e)

    output_file = os.path.join(args.output_path, "ear_phenotyping.xlsx")
    wb = load_workbook(output_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        label = row[0].value
        projection_pic_path = os.path.join(args.projection_image_folder, f'{label}.png')
        phenotype_prediction = phenotype_classification(projection_pic_path, model1)
        kernel_row_number_prediction = kernel_row_number_classification(projection_pic_path, model2)

        if phenotype_prediction == 0:
            for col in range(2, ws.max_column + 1): 
                ws.cell(row=row[0].row, column=col, value='NA')
        else:
            if kernel_row_number_prediction == 0:
                for col in range(7, 12): 
                    ws.cell(row=row[0].row, column=col, value='NA') 

    wb.save(output_file)


